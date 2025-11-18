# SPDX-FileCopyrightText: 2025 robot-visual-perception
#
# SPDX-License-Identifier: MIT

"""
SBOM Generator for robot-visual-perception project.

Generates:
- SBOM in CycloneDX JSON format (sbom.json)
- CSV with first-order dependencies (sbom-dependencies.csv)
- Optional: Updates planning-documents.xlsx in Deliverables/sprint-XX/
"""

import argparse
import csv
import json
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.request import urlopen
from urllib.error import HTTPError, URLError


def run_command(cmd: List[str], cwd: Optional[Path] = None) -> str:
    """Run shell command and return output."""
    try:
        result = subprocess.run(
            cmd, cwd=cwd, capture_output=True, text=True, check=True
        )
        return result.stdout.strip()
    except subprocess.CalledProcessError as e:
        print(f"Error running command {' '.join(cmd)}: {e.stderr}", file=sys.stderr)
        sys.exit(1)


def get_project_root() -> Path:
    """Get project root directory."""
    return Path(__file__).parent.parent


def get_npm_first_order_deps(package_json_path: Path) -> List[Dict[str, Any]]:
    """Extract first-order npm dependencies from package.json."""
    if not package_json_path.exists():
        return []

    with open(package_json_path, "r") as f:
        data = json.load(f)

    deps = []

    # Production dependencies
    for name, version in data.get("dependencies", {}).items():
        version_clean = version.lstrip("^~>=<")
        deps.append({
            "name": name,
            "version": version_clean,
            "ecosystem": "npm",
            "is_dev": False
        })

    # Dev dependencies
    for name, version in data.get("devDependencies", {}).items():
        version_clean = version.lstrip("^~>=<")
        deps.append({
            "name": name,
            "version": version_clean,
            "ecosystem": "npm",
            "is_dev": True
        })

    return deps


def get_python_first_order_deps(requirements_path: Path) -> List[Dict[str, Any]]:
    """Extract first-order Python dependencies from requirements.txt."""
    if not requirements_path.exists():
        return []

    deps = []
    with open(requirements_path, "r") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue

            # Parse package==version
            match = re.match(r"^([a-zA-Z0-9\-_.]+)\s*==\s*([0-9.]+)", line)
            if match:
                name, version = match.groups()
                deps.append({
                    "name": name,
                    "version": version,
                    "ecosystem": "pypi",
                    "is_dev": False
                })

    return deps


def fetch_license_from_npm(name: str, version: str) -> str:
    """Fetch license from npm registry."""
    from urllib.parse import quote

    encoded_name = quote(name, safe="@/")
    url = f"https://registry.npmjs.org/{encoded_name}"

    try:
        with urlopen(url, timeout=10) as response:
            data = json.load(response)

        # Try version-specific license first
        version_data = data.get("versions", {}).get(version, {})
        license_value = version_data.get("license")

        # Fallback to latest
        if not license_value:
            license_value = data.get("license")

        # Handle different license formats
        if isinstance(license_value, dict):
            license_type = license_value.get("type", "NOASSERTION")
            return license_type if license_type else "NOASSERTION"
        elif isinstance(license_value, str):
            return license_value if license_value else "NOASSERTION"

        return "NOASSERTION"
    except (HTTPError, URLError, TimeoutError, ValueError, KeyError) as e:
        print(f" (npm fetch error: {type(e).__name__})", end="")
        return "NOASSERTION"


def fetch_license_from_github(repo_url: str) -> str:
    """Try to fetch license from GitHub repository."""
    from urllib.parse import urlparse
    import re

    try:
        # Extract owner/repo from GitHub URL
        parsed = urlparse(repo_url)
        if "github.com" not in parsed.netloc.lower():
            return "NOASSERTION"

        match = re.search(r"github\.com/([^/]+)/([^/]+)", repo_url, re.IGNORECASE)
        if not match:
            return "NOASSERTION"

        owner, repo = match.groups()
        repo = repo.rstrip(".git")

        # Use GitHub API to get license
        api_url = f"https://api.github.com/repos/{owner}/{repo}"
        with urlopen(api_url, timeout=10) as response:
            data = json.load(response)
            license_data = data.get("license")
            if license_data and isinstance(license_data, dict):
                spdx_id = license_data.get("spdx_id")
                if spdx_id and spdx_id != "NOASSERTION":
                    return spdx_id

        return "NOASSERTION"
    except (HTTPError, URLError, TimeoutError, ValueError, KeyError):
        return "NOASSERTION"


def fetch_license_from_pypi(name: str, version: str) -> str:
    """Fetch license from PyPI."""
    from urllib.parse import quote

    encoded_name = quote(name)
    url = f"https://pypi.org/pypi/{encoded_name}/{version}/json"

    try:
        with urlopen(url, timeout=10) as response:
            data = json.load(response)

        info = data.get("info", {})

        # Try license field first
        license_field = info.get("license")
        if license_field:
            license_value = str(license_field).strip()

            # Skip empty or unknown values
            if not license_value or license_value.upper() in ("UNKNOWN", "NONE", "NULL"):
                license_value = None
            # Truncate very long licenses (e.g., full license text)
            elif len(license_value) > 100:
                # Try to extract SPDX identifier from long text
                upper_text = license_value.upper()
                if "BSD-3-CLAUSE" in upper_text or ("BSD" in upper_text and "3-CLAUSE" in upper_text):
                    return "BSD-3-Clause"
                elif "BSD-2-CLAUSE" in upper_text or ("BSD" in upper_text and "2-CLAUSE" in upper_text):
                    return "BSD-2-Clause"
                elif "BSD" in upper_text:
                    return "BSD-3-Clause"
                elif "MIT" in upper_text:
                    return "MIT"
                elif "APACHE-2.0" in upper_text or "APACHE 2.0" in upper_text:
                    return "Apache-2.0"
                elif "APACHE" in upper_text:
                    return "Apache-2.0"
                elif "GPL-3" in upper_text or "GPLV3" in upper_text:
                    return "GPL-3.0"
                elif "GPL-2" in upper_text or "GPLV2" in upper_text:
                    return "GPL-2.0"
                elif "GPL" in upper_text:
                    return "GPL"
                elif "LGPL" in upper_text:
                    return "LGPL"
                else:
                    license_value = None

            if license_value:
                return license_value

        # Fallback to classifiers
        classifiers = info.get("classifiers", [])
        for classifier in reversed(classifiers):
            if classifier.startswith("License ::"):
                parts = [p.strip() for p in classifier.split("::") if p.strip()]
                if len(parts) > 1 and parts[-1] not in ("OSI Approved", "License"):
                    # Convert common classifier names to SPDX
                    license_name = parts[-1]
                    if "MIT" in license_name:
                        return "MIT"
                    elif "Apache" in license_name and "2.0" in license_name:
                        return "Apache-2.0"
                    elif "BSD" in license_name and "3-Clause" in license_name:
                        return "BSD-3-Clause"
                    elif "BSD" in license_name and "2-Clause" in license_name:
                        return "BSD-2-Clause"
                    elif "GPL" in license_name:
                        return license_name
                    return license_name

        # Last resort: try to get license from project's GitHub repo
        project_urls = info.get("project_urls", {})
        for url_type, url in project_urls.items():
            if url and ("github.com" in url.lower() or "source" in url_type.lower() or "repository" in url_type.lower()):
                github_license = fetch_license_from_github(url)
                if github_license != "NOASSERTION":
                    return github_license

        # Also check home_page field
        home_page = info.get("home_page")
        if home_page and "github.com" in home_page.lower():
            github_license = fetch_license_from_github(home_page)
            if github_license != "NOASSERTION":
                return github_license

        return "NOASSERTION"
    except (HTTPError, URLError, TimeoutError, ValueError, KeyError) as e:
        print(f" (pypi fetch error: {type(e).__name__})", end="")
        return "NOASSERTION"


def enrich_with_licenses(deps: List[Dict[str, Any]]) -> None:
    """Enrich dependencies with license information."""
    for dep in deps:
        ecosystem = dep["ecosystem"]
        name = dep["name"]
        version = dep["version"]

        print(f"  Fetching license for {name}@{version}...", end="")

        if ecosystem == "npm":
            dep["license"] = fetch_license_from_npm(name, version)
        elif ecosystem == "pypi":
            dep["license"] = fetch_license_from_pypi(name, version)
        else:
            dep["license"] = "NOASSERTION"

        print(f" {dep['license']}")


def generate_cyclonedx_sbom(all_deps: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Generate CycloneDX SBOM."""
    timestamp = datetime.utcnow().isoformat() + "Z"

    sbom = {
        "bomFormat": "CycloneDX",
        "specVersion": "1.6",
        "version": 1,
        "metadata": {
            "timestamp": timestamp,
            "tools": [{
                "name": "robot-visual-perception-sbom-generator",
                "version": "2.0.0"
            }],
            "component": {
                "type": "application",
                "name": "robot-visual-perception",
                "version": "0.1.0"
            }
        },
        "components": []
    }

    for dep in all_deps:
        component = {
            "type": "library",
            "name": dep["name"],
            "version": dep["version"],
        }

        # Add PURL
        if dep["ecosystem"] == "npm":
            component["purl"] = f"pkg:npm/{dep['name']}@{dep['version']}"
        elif dep["ecosystem"] == "pypi":
            component["purl"] = f"pkg:pypi/{dep['name']}@{dep['version']}"

        # Add license if available
        if dep.get("license") and dep["license"] != "NOASSERTION":
            component["licenses"] = [{
                "license": {
                    "id": dep["license"]
                }
            }]

        sbom["components"].append(component)

    return sbom


def generate_csv_for_planning_doc(all_deps: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    """Generate CSV data matching the planning document format."""
    csv_data = []

    # Context mapping
    def get_context(dep: Dict[str, Any]) -> str:
        if dep["ecosystem"] == "npm":
            return "Frontend (React UI)"
        elif dep["ecosystem"] == "pypi":
            if dep.get("is_dev"):
                return "Backend Dev Dependencies"
            else:
                return "Backend (FastAPI API)"
        return "Other"

    # Sort: Frontend first, then Backend, then Dev
    sorted_deps = sorted(all_deps, key=lambda d: (
        0 if d["ecosystem"] == "npm" and not d.get("is_dev") else
        1 if d["ecosystem"] == "pypi" and not d.get("is_dev") else
        2,
        d["name"]
    ))

    for idx, dep in enumerate(sorted_deps, start=1):
        context = get_context(dep)

        # Format name with ecosystem prefix
        if dep["ecosystem"] == "npm":
            formatted_name = f"npm:{dep['name']}"
        elif dep["ecosystem"] == "pypi":
            formatted_name = f"pypi:{dep['name']}"
        else:
            formatted_name = dep["name"]

        csv_data.append({
            "#": str(idx),
            "Context": context,
            "Name": formatted_name,
            "Version": dep["version"],
            "License": dep.get("license", "NOASSERTION"),
            "Comment": ""  # Empty for automated entries
        })

    return csv_data


def write_csv(csv_data: List[Dict[str, str]], output_path: Path) -> None:
    """Write CSV file."""
    fieldnames = ["#", "Context", "Name", "Version", "License", "Comment"]

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(csv_data)


def get_latest_sprint_number() -> Optional[int]:
    """Get the latest sprint number from Git tags."""
    try:
        tag = run_command(["git", "describe", "--tags", "--abbrev=0"])
        match = re.search(r"sprint-(\d+)", tag)
        if match:
            return int(match.group(1))
    except Exception:
        pass

    # Fallback: Find highest sprint folder
    root = get_project_root()
    deliverables = root / "Deliverables"
    if not deliverables.exists():
        return None

    sprint_folders = [
        int(f.name.replace("sprint-", ""))
        for f in deliverables.iterdir()
        if f.is_dir() and f.name.startswith("sprint-")
    ]
    return max(sprint_folders) if sprint_folders else None


def prepare_sprint_folder(current_sprint: int) -> Path:
    """Ensure next sprint folder exists and has planning-documents.xlsx."""
    root = get_project_root()
    deliverables = root / "Deliverables"
    next_sprint = current_sprint + 1

    next_sprint_folder = deliverables / f"sprint-{next_sprint:02d}"
    current_sprint_folder = deliverables / f"sprint-{current_sprint:02d}"

    if not next_sprint_folder.exists():
        print(f"Creating new sprint folder: {next_sprint_folder.name}")
        next_sprint_folder.mkdir(parents=True)

    import shutil

    def copy_if_missing(source: Path, target: Path, label: str) -> None:
        if target.exists():
            return
        if source.exists():
            shutil.copy2(source, target)
            print(f"Copied {label} from sprint-{current_sprint:02d}")
        else:
            print(
                f"Warning: {source} not found, cannot copy to new sprint",
                file=sys.stderr,
            )

    copy_if_missing(
        current_sprint_folder / "planning-documents.xlsx",
        next_sprint_folder / "planning-documents.xlsx",
        "planning-documents.xlsx",
    )
    copy_if_missing(
        current_sprint_folder / "planning-documents.xlsx.license",
        next_sprint_folder / "planning-documents.xlsx.license",
        "planning-documents.xlsx.license",
    )

    return next_sprint_folder


def update_excel(csv_data: List[Dict[str, str]], sprint_folder: Path) -> None:
    """Update planning-documents.xlsx with new BOM data."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except ImportError:
        print(
            "Error: openpyxl not installed. Run: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    excel_path = sprint_folder / "planning-documents.xlsx"
    if not excel_path.exists():
        print(f"Error: {excel_path} not found", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(excel_path)

    # Find or create 'Bill of Materials' sheet
    sheet_name = "Bill of Materials"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)

    ws = wb.create_sheet(sheet_name)

    # Write headers
    headers = list(csv_data[0].keys())
    ws.append(headers)

    header_fill = PatternFill(
        fill_type="solid",
        start_color="FFD9D9D9",
        end_color="FFD9D9D9",
    )
    for cell in ws[1]:
        cell.fill = header_fill

    # Write data
    for row in csv_data:
        ws.append(list(row.values()))

    # Set column widths for better readability
    # Column widths: #, Context, Name, Version, License, Comment
    column_widths = [5, 30, 40, 15, 25, 40]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # Save
    wb.save(excel_path)
    print(f"Updated {excel_path}")


def main():
    parser = argparse.ArgumentParser(description="Generate SBOM for the project")
    parser.add_argument(
        "--update-excel",
        action="store_true",
        help="Update planning-documents.xlsx in Deliverables folder",
    )
    parser.add_argument(
        "--check",
        action="store_true",
        help="Check if SBOM files are up-to-date (exit 1 if not)",
    )
    args = parser.parse_args()

    root = get_project_root()
    sbom_path = root / "sbom.json"
    csv_path = root / "sbom-dependencies.csv"

    # Check mode
    if args.check:
        if not sbom_path.exists() or not csv_path.exists():
            print("Error: SBOM files missing. Run 'make sbom' to generate.")
            sys.exit(1)

        # Check if dependency files are newer than SBOM
        dependency_files = [
            root / "src" / "backend" / "requirements.txt",
            root / "src" / "backend" / "requirements-dev.txt",
            root / "src" / "frontend" / "package.json",
        ]

        existing_mtimes = [
            path.stat().st_mtime for path in dependency_files if path.exists()
        ]
        if not existing_mtimes:
            print("Error: No dependency files found.", file=sys.stderr)
            sys.exit(1)

        req_time = max(existing_mtimes)
        sbom_time = min(sbom_path.stat().st_mtime, csv_path.stat().st_mtime)

        if req_time > sbom_time:
            print("Error: Dependencies changed but SBOM not updated. Run 'make sbom'.")
            sys.exit(1)

        print("SBOM is up-to-date.")
        return

    # Collect first-order dependencies
    print("Parsing first-order dependencies...")

    all_deps = []

    # Frontend dependencies
    frontend_deps = get_npm_first_order_deps(root / "src" / "frontend" / "package.json")
    all_deps.extend(frontend_deps)
    print(f"Found {len(frontend_deps)} frontend dependencies")

    # Backend production dependencies
    backend_deps = get_python_first_order_deps(root / "src" / "backend" / "requirements.txt")
    all_deps.extend(backend_deps)
    print(f"Found {len(backend_deps)} backend production dependencies")

    # Backend dev dependencies
    backend_dev_deps = get_python_first_order_deps(root / "src" / "backend" / "requirements-dev.txt")
    for dep in backend_dev_deps:
        dep["is_dev"] = True
    all_deps.extend(backend_dev_deps)
    print(f"Found {len(backend_dev_deps)} backend dev dependencies")

    print(f"\nTotal first-order dependencies: {len(all_deps)}")

    # Enrich with license information
    print("\nFetching license information...")
    enrich_with_licenses(all_deps)

    # Generate SBOM
    print("\nGenerating CycloneDX SBOM...")
    sbom = generate_cyclonedx_sbom(all_deps)

    with open(sbom_path, "w") as f:
        json.dump(sbom, f, indent=2)
    print(f"Generated: {sbom_path}")

    # Generate CSV
    print("Generating CSV for planning document...")
    csv_data = generate_csv_for_planning_doc(all_deps)
    write_csv(csv_data, csv_path)
    print(f"Generated: {csv_path} ({len(csv_data)} dependencies)")

    # Update Excel if requested
    if args.update_excel:
        print("\nUpdating Excel...")
        sprint_num = get_latest_sprint_number()

        if sprint_num is None:
            print("Error: Cannot determine current sprint number", file=sys.stderr)
            sys.exit(1)

        sprint_folder = prepare_sprint_folder(sprint_num)
        update_excel(csv_data, sprint_folder)

    print("\nDone!")


if __name__ == "__main__":
    main()